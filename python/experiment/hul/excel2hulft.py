import openpyxl
import sys
import os

class HulftRcvParser:
    def __init__(self, file_path):
        self.__set_header()
        lines = self.__readlines(file_path)
        self.__load(lines)

    def __set_header(self):
        self.__headers = [
            'rcvfile',
            'filename',
            'owner',
            'group',
            'perm',
            'transmode',
            'abnormal',
            'rcvtype',
            'jobid',
            'comment',
            'grpid',
            'ejobid',
            'genctl',
            'password',
            'codeset',
            'jobwait',
            'genmngno',
            'dataverify'
        ]

    def __readlines(self, file_path):
        with open(file_path, mode='r') as f:
            lines = f.readlines()
        return lines

    def __load(self, lines):
        self.hulft_rcvs = []
        self.__params = {}
        for line in lines:
            self.__parse(line.strip())

    def __parse(self, line):
        if line == 'END':
            self.hulft_rcvs.append(HulftRcv(self.__params))
            self.__params = {}
            return
        if line == '':
            #何もしない
            return
        if line[0] == '#':
            #何もしない
            return

        splited_strings = line.split('=')
        key = splited_strings[0].lower()
        if key in self.__headers:
            self.__params[key] = splited_strings[1]



class HulftRcv:
    def __init__(self, params):
        self.rcvfile = params['rcvfile']
        self.filename = params['filename']
        self.owner = params['owner']
        self.group = params['group']
        self.perm = params['perm']
        self.transmode = params['transmode']
        self.abnormal = params['abnormal']
        self.rcvtype = params['rcvtype']
        self.jobid = params['jobid']
        self.comment = params['comment']
        self.grpid = params['grpid']
        self.ejobid = params['ejobid']
        self.genctl = params['genctl']
        self.password = params['password']
        self.codeset = params['codeset']
        self.jobwait = params['jobwait']
        self.genmngno = params['genmngno']
        self.dataverify = params['dataverify']

class ExcelWriter:
    def __init__(self, hulftrcv):
        self.__hulft_rcvs = hulftrcv.hulft_rcvs
        self.__set_excel()
        self.__set_header()

    def __set_excel(self):
        self.__output_file_name = "hulft.xlsx"
        self.__sheet_name = "Sheet1"

        self.__set_output_file()
        self.__set_sheet()

    def __set_output_file(self):
        self.__output_file = openpyxl.load_workbook(self.__output_file_name)

    def __set_sheet(self):
        self.__sheet = self.__output_file[self.__sheet_name]

    def __set_header(self):
        self.__headers = [
            'rcvfile',
            'filename',
            'owner',
            'group',
            'perm',
            'transmode',
            'abnormal',
            'rcvtype',
            'jobid',
            'comment',
            'grpid',
            'ejobid',
            'genctl',
            'password',
            'codeset',
            'jobwait',
            'genmngno',
            'dataverify'
        ]

    def run(self):
        self.__write_header()

        for item in self.hulft_rcvs:
            self.__write_body(item)

        self.__output_file.save(self.__output_file_name)

    def __write_header(self):
        for i, header in enumerate(self.__headers):
            self.__sheet.cell(row=1, column=i+1).value = header

    def __write_body(self):
        for i, hulftrcv in enumerate(self.__hulft_rcvs):
            for j, header in enumerate(self.__headers):
                self.__sheet.cell(row=i+1, column=j+1).value = getattr(hulftrcv, header)


def main():
    args = sys.argv
    if len(args) < 2:
        print("ファイル名を引数に指定してください")
        return

    file = args[1]
    if not os.path.exists(file):
        print(file)
        print("ファイルパスが正しくありません")
        return

    hulrcv = HulftRcvParser(file)
    # for item in hulrcv.hulft_rcvs:
    #     print(item.rcvfile)
    writer = ExcelWriter(hulrcv)
    writer.run()

main()
