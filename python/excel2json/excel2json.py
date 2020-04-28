import openpyxl
import json


def main():
    settings = json.load(open('settings.json', 'r'))

    book = openpyxl.load_workbook(settings['dataFile'])
    sheet = book[settings['dataSheet']]
    output = settings['outputFile']
    output_type = settings['outputFileType']

    # Header
    with open(output, mode='w') as f:
        f.write("[")

    columns = set_columns(sheet)

    row = 2
    while check(sheet, row, 1):
        # add comma
        if row != 2:
            with open(output, mode='a') as f:
                f.write(",")

        data = {}
        for index, column in enumerate(columns):
            data[column] = sheet.cell(row=row, column=index+1).value

        write_data(output, data, output_type)

        row += 1

    # Footer
    with open(output, mode='a') as f:
        f.write("]")


def set_columns(sheet):
    columns = []
    column = 1
    while check(sheet, 1, column):
        columns.append(sheet.cell(row=1, column=column).value)
        column += 1
    return columns


def check(sheet, row, column):
    data = sheet.cell(row=row, column=column).value
    if data != '' and data != None:
        return True
    return False


def write_data(output, data, output_type):
    if output_type == 'json':
        with open(output, mode='a') as f:
            f.write(json.dumps(data, sort_keys=False,
                               ensure_ascii=False, indent=4))
    elif output_type == 'php':
        with open(output, mode='a') as f:
            f.write('[\n    ')
            print(data)
            for key, value in data.items():
                f.write('\"'+str(key)+'\"'+' => ')
                if isinstance(value, (int, float)):
                    f.write(str(value))
                else:
                    f.write('\"'+str(value)+'\"')
                if key != len(data):
                    f.write(',')
            f.write('\n]')


main()
